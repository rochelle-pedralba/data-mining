import pandas as pd
import os
from openpyxl import load_workbook

def save_excel(rules, country):
  rules['antecedents'] = rules['antecedents'].apply(lambda x: ', '.join(list(x)))
  rules['consequents'] = rules['consequents'].apply(lambda x: ', '.join(list(x)))

  excel_path = "../results/association_rules.xlsx"

  if os.path.exists(excel_path):
    book = load_workbook(excel_path)

    if country in book.sheetnames:
      del book[country]
    book.save(excel_path)
    book.close()

  with pd.ExcelWriter("../results/association_rules.xlsx", engine='openpyxl', mode='a') as writer:
    rules.to_excel(writer, sheet_name=country, index=False)